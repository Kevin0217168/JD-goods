# -*- coding:utf8 -*-
from os import remove, path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


class XlsxSaver:
    """
    一个将DataFrame转换成格式化excel的工具
    """
    
    def __init__(self, df_in, filename='a.xlsx', sheet_name='Sheet1'):
        """
        df_in : 从一个DataFrame对象获取表格内容
        filename : 文件名
        sheet_name : 表名
        """
        self.filename = filename  # 保存的xlsx文件的名字
        self.user_def = []  # 储存由用户自定义的列的列名，这些列不再参与自动计算列宽
        if path.exists(filename):
            # 如果文件存在，就直接打开，添加Sheet
            self.wb = load_workbook(filename)
            self.sheet = self.wb.create_sheet(sheet_name)
        else:
            # 如果文件不存在，就创建表格
            self.wb = Workbook()
            self.sheet = self.wb.active
            self.sheet.title = sheet_name
        # 将df的内容复制给sheet
        self.df = df_in.copy()
        self.sheet.append(list(self.df.columns))
        for row in range(0, len(list(self.df.index))):
            for col in range(0, len(list(self.df.columns))):
                self.sheet.cell(row + 2, col + 1).value = self.df.iloc[row, col]  # 注意：sheet行列从1开始计数
    
    def remove_file(self):
        remove(self.filename)
    
    def set_sheet_name(self, sheet_name):
        self.sheet.title = sheet_name
    
    def set_filename(self, filename):
        self.filename = filename
    
    @staticmethod
    def get_maxlength(series_in, col):
        """
        获取一个类型为object的Series中的最大占位长度，用于确定导出的xlsx文件的列宽
        col : 表头，也参与比较，解决有时候表头过长的问题
        """
        series = series_in.fillna('-')  # 填充空值，防止出现nan
        str_list = list(series)
        len_list = []
        for elem in str_list + [col]:
            elem_split = list(elem)
            length = 0
            for c in elem_split:
                if ord(c) <= 256:
                    length += 1
                else:
                    length += 2
            len_list.append(length)
        return max(len_list)
    
    def __auto_width(self):
        cols_list = list(self.df.columns)  # 获取列名
        for i in range(0, len(cols_list)):
            col = cols_list[i]
            if col in self.user_def:
                continue
            self.sheet.cell(1, i + 1).font = Font(bold=True)  # 加粗表头
            letter = chr(i + 65)  # 由ASCII值获得对应的列字母
            max_len = self.get_maxlength(self.df[col].astype(str), col)
            if max_len <= 12:
                self.sheet.column_dimensions[letter].width = 12
            elif max_len <= 50:
                self.sheet.column_dimensions[letter].width = max_len + 2
            else:
                self.sheet.column_dimensions[letter].width = 50
                for cell in self.sheet[letter]:
                    cell.alignment = Alignment(wrap_text=True)
    
    def set_width(self, col_name, width):
        # 提供调整列宽的接口
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        self.sheet.column_dimensions[letter].width = width
        self.user_def.append(col_name)
    
    def set_color(self, col_name, color, rule):
        # 提供设置颜色的接口，rule:规则函数
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        for cell in self.sheet[letter]:
            if rule(cell.value):
                cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
    
    def set_center_alignment(self, col_name):
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        for cell in self.sheet[letter]:
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
    
    def save(self):
        # 自动调整列宽，并保存
        self.__auto_width()
        self.wb.save(self.filename)
    
    def set_merge(self, col_name):
        self.user_def.append(col_name)  # 设置为自定义列
        # 设置一列合并单元格
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        i = 1
        while True:
            if i >= self.sheet.max_row:
                # 结束条件：单元格到底
                break
            cell = self.sheet[letter + str(i)]
            j = i + 1  # 第一步指向下一个单元格
            while True:
                # 这个过程对j进行试探，最终j指向的单元格是与i连续相同的最后一个
                cell_next = self.sheet[letter + str(j)]
                if cell_next.value != cell.value:
                    j -= 1
                    break
                else:
                    j += 1
                if j > self.sheet.max_row:
                    j -= 1
                    break
            if j - i >= 1 and cell.value != '' and cell.value:
                # 如果有连续两格以上的单元格内容相同，进行融合
                msg = '%s%d:%s%d' % (letter, i, letter, j)
                self.sheet.merge_cells(msg)
            # 控制一下格式
            self.sheet[letter + str(i)].alignment = Alignment(horizontal='center',
                                                              vertical='top',
                                                              wrap_text=True)
            i = j + 1  # 继续指向下个单元格
